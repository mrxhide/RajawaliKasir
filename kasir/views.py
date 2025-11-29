from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.utils import timezone
from django.db import models
from django.db.models import Sum, Count, F
from django.db.models.functions import TruncDate, ExtractHour
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.urls import reverse
from django.contrib.auth import logout
from django.middleware.csrf import get_token


import uuid
from decimal import Decimal
from io import BytesIO

import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4

import qrcode
import base64
from django.conf import settings

from .models import Product, Sale, SaleItem, StockMovement
from .forms import ProductForm, StokMasukForm, StokKeluarForm
from datetime import datetime
from django.utils.timezone import make_aware
from django.db import transaction
from openpyxl.utils import get_column_letter


# =====================================================
# =============            DASHBOARD        ============
# =====================================================

from django.utils.timezone import localdate, localtime

@login_required
def dashboard(request):

    # Ambil tanggal lokal WIB
    today = localdate()

    # Range 7 hari WIB
    seven_days_ago = today - timezone.timedelta(days=6)

    # -------------------------------
    # TOTAL HARI INI (menggunakan WIB)
    # -------------------------------
    total_penjualan_hari_ini = (
        Sale.objects
        .filter(date_time__date=today)
        .aggregate(total=Sum("total"))["total"] or 0
    )

    total_transaksi_hari_ini = Sale.objects.filter(
        date_time__date=today
    ).count()

    # -------------------------------
    # STOK MENIPIS
    # -------------------------------
    stok_menipis = Product.objects.filter(stock__lte=models.F("min_stock"))

    # -------------------------------
    # PENJUALAN 7 HARI TERAKHIR
    # -------------------------------
    daily_qs = (
        Sale.objects
        .filter(date_time__date__gte=seven_days_ago)
        .annotate(day=TruncDate("date_time"))
        .values("day")
        .annotate(total=Sum("total"))
        .order_by("day")
    )

    daily_labels = [d["day"].strftime("%d-%m") for d in daily_qs]
    daily_values = [float(d["total"]) for d in daily_qs]

    # -------------------------------
    # PRODUK TERLARIS
    # -------------------------------
    top_products_qs = (
        SaleItem.objects
        .values("product__name")
        .annotate(total_qty=Sum("qty"))
        .order_by("-total_qty")[:5]
    )

    top_products_labels = [p["product__name"] for p in top_products_qs]
    top_products_values = [int(p["total_qty"]) for p in top_products_qs]

    # -------------------------------
    # JAM TERSIBUK HARI INI
    # -------------------------------
    hourly_qs = (
        Sale.objects
        .filter(date_time__date=today)
        .annotate(hour=ExtractHour("date_time"))
        .values("hour")
        .annotate(count=Count("id"))
        .order_by("hour")
    )

    hourly_labels = [f"{h['hour']:02d}:00" for h in hourly_qs]
    hourly_values = [int(h["count"]) for h in hourly_qs]

    # -------------------------------
    # STOK MENIPIS UNTUK CHART
    # -------------------------------
    low_stock_qs = Product.objects.filter(stock__lte=models.F("min_stock"))
    low_stock_labels = [p.name for p in low_stock_qs]
    low_stock_values = [p.stock for p in low_stock_qs]

    # -------------------------------
    # FINAL CONTEXT
    # -------------------------------
    context = {
        "total_penjualan_hari_ini": total_penjualan_hari_ini,
        "total_transaksi_hari_ini": total_transaksi_hari_ini,

        "daily_labels": daily_labels,
        "daily_values": daily_values,

        "top_products_labels": top_products_labels,
        "top_products_values": top_products_values,

        "hourly_labels": hourly_labels,
        "hourly_values": hourly_values,

        "low_stock_labels": low_stock_labels,
        "low_stock_values": low_stock_values,

        "stok_menipis": stok_menipis,
    }

    return render(request, "kasir/dashboard.html", context)

# =====================================================
# =============            POS SYSTEM       ============
# =====================================================

# -------------------------------
# POS PAGE
# -------------------------------
def pos(request):
    pos_success = request.session.get("pos_success", None)
    return render(request, "kasir/pos.html", {"pos_success": pos_success})



# -------------------------------
# CART SESSION HELPER
# -------------------------------
def get_cart(request):
    return request.session.get("cart", {})

def save_cart(request, cart):
    request.session["cart"] = cart
    request.session.modified = True


# -------------------------------
# PRODUCT LIST (GRID) – HTMX
# -------------------------------
def pos_items(request):
    q = request.GET.get("q", "")

    products = Product.objects.filter(is_active=True)

    if q:
        products = products.filter(name__icontains=q)

    products = products.order_by("name")

    return render(request, "kasir/pos_items.html", {
        "products": products
    })


# -------------------------------
# CART – HTMX
# -------------------------------
def pos_cart(request):
    cart = get_cart(request)
    cart_items = []
    total = 0

    for pid, qty in cart.items():
        try:
            product = Product.objects.get(id=pid)
        except Product.DoesNotExist:
            continue

        line_total = product.sell_price * qty
        total += line_total

        cart_items.append({
            "id": pid,
            "product": product,
            "qty": qty,
            "total": line_total,
        })

    return render(request, "kasir/pos_cart.html", {
        "cart": cart_items,
        "total": total,
    })


# -------------------------------
# ADD TO CART
# -------------------------------
def cart_add(request, product_id):
    product = Product.objects.get(id=product_id)

    # Stok habis
    if product.stock <= 0:
        return render(request, "kasir/error_sound.html", {
            "message": f"Stok {product.name} habis!"
        })

    cart = get_cart(request)
    pid = str(product_id)

    # Jika qty ingin naik tapi melewati stok → error
    if cart.get(pid, 0) >= product.stock:
        return render(request, "kasir/error_sound.html", {
            "message": f"Stok {product.name} tidak mencukupi!"
        })

    cart[pid] = cart.get(pid, 0) + 1
    save_cart(request, cart)

    return pos_cart(request)


# -------------------------------
# INCREASE QTY
# -------------------------------
def cart_increase(request, item_id):
    cart = get_cart(request)
    pid = str(item_id)

    product = Product.objects.get(id=pid)

    if cart.get(pid, 0) >= product.stock:
        return render(request, "kasir/error_sound.html", {
            "message": f"Stok {product.name} tidak mencukupi!"
        })

    cart[pid] += 1
    save_cart(request, cart)

    return pos_cart(request)


# -------------------------------
# DECREASE QTY
# -------------------------------
def cart_decrease(request, item_id):
    cart = get_cart(request)
    pid = str(item_id)

    if pid in cart:
        if cart[pid] > 1:
            cart[pid] -= 1
        else:
            del cart[pid]

    save_cart(request, cart)
    return pos_cart(request)


# -------------------------------
# REMOVE FROM CART
# -------------------------------
def cart_remove(request, item_id):
    cart = get_cart(request)
    pid = str(item_id)

    if pid in cart:
        del cart[pid]

    save_cart(request, cart)
    return pos_cart(request)


# -------------------------------
# CHECKOUT PAGE
# -------------------------------
def pos_checkout(request):

    # Generate QRIS
    qris_data = settings.QRIS_CONTENT
    qr = qrcode.make(qris_data)
    buffer = BytesIO()
    qr.save(buffer, format="PNG")
    qris_base64 = base64.b64encode(buffer.getvalue()).decode()

    # Display modal
    if request.method == "GET":
        cart = get_cart(request)
        total = 0
        for pid, qty in cart.items():
            product = Product.objects.get(id=pid)
            total += product.sell_price * qty

        return render(request, "kasir/pos_checkout.html", {
            "total": total,
            "qris_image": qris_base64,
        })

    # Process checkout
    cart = get_cart(request)
    if not cart:
        return HttpResponse("Keranjang kosong!")

    total = 0
    items = []

    for pid, qty in cart.items():
        product = Product.objects.get(id=pid)

        if qty > product.stock:
            return render(request, "kasir/error_sound.html", {
                "message": f"Stok {product.name} tidak cukup!"
            })

        line_total = product.sell_price * qty
        total += line_total
        items.append((product, qty, line_total))

    payment_method = request.POST.get("payment_method")

    if payment_method == "qris":
        paid = total
        change = 0
    else:
        paid = Decimal(request.POST.get("paid"))
        change = paid - total

    invoice_no = "RJW-" + uuid.uuid4().hex[:8].upper()

    sale = Sale.objects.create(
        invoice_no=invoice_no,
        cashier=request.user,
        subtotal=total,
        discount=0,
        total=total,
        payment_method=payment_method,
        paid_amount=paid,
        change_amount=change,
    )

    for product, qty, line_total in items:
        SaleItem.objects.create(
            sale=sale,
            product=product,
            qty=qty,
            price=product.sell_price,
            total=line_total,
        )

        product.stock -= qty
        product.save()

        StockMovement.objects.create(
            product=product,
            movement_type="SALE",
            quantity=-qty,
            note=f"Penjualan {invoice_no}",
            user=request.user,
        )

    save_cart(request, {})

    request.session["pos_success"] = invoice_no
    request.session.modified = True

    response = HttpResponse()
    response["HX-Redirect"] = f"/pos/struk/{invoice_no}/"
    return response


# -------------------------------
# STRUK PAGE
# -------------------------------
def pos_struk(request, invoice):
    sale = Sale.objects.get(invoice_no=invoice)
    items = SaleItem.objects.filter(sale=sale)

    return render(request, "kasir/struk.html", {
        "sale": sale,
        "items": items,
    })


# -------------------------------
# REMOVE pos_success SESSION
# -------------------------------
def remove_pos_success(request):
    request.session.pop("pos_success", None)
    request.session.modified = True
    return HttpResponse("OK")




# =====================================================
# =============            PRODUK           ============
# =====================================================

def product_list(request):
    products = Product.objects.all().order_by("name")
    return render(request, "kasir/product_list.html", {"products": products})


def product_add(request):
    if request.method == "POST":
        form = ProductForm(request.POST, request.FILES)  # <-- WAJIB ADA request.FILES
        if form.is_valid():
            form.save()
            messages.success(request, "Produk berhasil ditambahkan!")
            return redirect("kasir:product_list")
    else:
        form = ProductForm()

    return render(
        request,
        "kasir/product_form.html",
        {"form": form, "title": "Tambah Produk"},
    )

def cart_increase(request, item_id):
    cart = get_cart(request)
    item_id = str(item_id)

    if item_id in cart:
        product = Product.objects.get(id=item_id)

        if cart[item_id] >= product.stock:
            return render(request, "kasir/error_stock_limit.html", {
                "message": f"Stok tidak cukup! Sisa stok hanya {product.stock}."
            })

        cart[item_id] += 1

    save_cart(request, cart)
    return pos_cart(request)


def cart_decrease(request, item_id):
    cart = get_cart(request)
    item_id = str(item_id)

    if item_id in cart:
        if cart[item_id] > 1:
            cart[item_id] -= 1
        else:
            del cart[item_id]

    save_cart(request, cart)
    return pos_cart(request)


def cart_decrease(request, item_id):
    cart = get_cart(request)
    item_id = str(item_id)

    if item_id in cart:
        if cart[item_id] > 1:
            cart[item_id] -= 1
        else:
            del cart[item_id]  # hapus jika qty jadi 0

    save_cart(request, cart)
    return pos_cart(request)


def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)

    if request.method == "POST":
        form = ProductForm(request.POST, request.FILES, instance=product)  # <-- WAJIB ADA request.FILES
        if form.is_valid():
            form.save()
            messages.success(request, "Produk berhasil diperbarui!")
            return redirect("kasir:product_list")
    else:
        form = ProductForm(instance=product)

    return render(
        request,
        "kasir/product_form.html",
        {"form": form, "title": "Edit Produk"},
    )



def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    product.is_active = False
    product.save()
    messages.warning(request, "Produk dinonaktifkan.")
    return redirect("kasir:product_list")


# =====================================================
# =============              STOK           ============
# =====================================================

def stok_masuk(request):
    if request.method == "POST":
        form = StokMasukForm(request.POST)
        if form.is_valid():
            product = form.cleaned_data["product"]
            quantity = form.cleaned_data["quantity"]
            note = form.cleaned_data["note"]

            product.stock += quantity
            product.save()

            StockMovement.objects.create(
                product=product,
                movement_type="IN",
                quantity=quantity,
                note=note or "Stok Masuk",
                user=request.user,
            )

            messages.success(request, "Stok berhasil ditambahkan.")
            return redirect("kasir:stok_masuk_list")
    else:
        form = StokMasukForm()

    return render(request, "kasir/stok_masuk.html", {"form": form})


def stok_masuk_list(request):
    movements = StockMovement.objects.filter(movement_type="IN").order_by(
        "-created_at"
    )
    return render(request, "kasir/stok_masuk_list.html", {"movements": movements})


def stok_keluar(request):
    if request.method == "POST":
        form = StokKeluarForm(request.POST)
        if form.is_valid():
            product = form.cleaned_data["product"]
            quantity = form.cleaned_data["quantity"]
            note = form.cleaned_data["note"]

            if quantity > product.stock:
                form.add_error("quantity", "Stok tidak mencukupi untuk dikurangi.")
            else:
                product.stock -= quantity
                product.save()

                StockMovement.objects.create(
                    product=product,
                    movement_type="OUT",
                    quantity=-quantity,
                    note=note or "Stok Keluar",
                    user=request.user,
                )

                messages.success(request, "Stok berhasil dikurangi.")
                return redirect("kasir:stok_keluar_list")
    else:
        form = StokKeluarForm()

    return render(request, "kasir/stok_keluar.html", {"form": form})


def stok_keluar_list(request):
    movements = StockMovement.objects.filter(movement_type="OUT").order_by(
        "-created_at"
    )
    return render(request, "kasir/stok_keluar_list.html", {"movements": movements})


# =====================================================
# =============            LAPORAN          ============
# =====================================================

def get_filtered_sales(request):
    from datetime import datetime

    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    cashier_id = request.GET.get("cashier")

    reset_detail = request.GET.get("reset_detail")
    reset_rekap = request.GET.get("reset_rekap")

    sales = Sale.objects.all().order_by("-date_time")

    # =====================================================
    # RESET DETAIL (hapus tampilan detail transaksi)
    # =====================================================
    if reset_detail:
        return {
            "sales": [],
            "rekap": [],
            "total_revenue": 0,
            "total_transactions": 0,
            "start_date": "",
            "end_date": "",
            "cashier_id": "",
            "reset_detail_msg": True,   # → untuk SweetAlert2
        }

    # =====================================================
    # RESET REKAP (hapus rekap per hari)
    # =====================================================
    if reset_rekap:
        return {
            "sales": sales,              # detail tetap tampil
            "rekap": [],                 # rekap dikosongkan
            "total_revenue": sales.aggregate(total=Sum("total"))["total"] or 0,
            "total_transactions": sales.count(),
            "start_date": "",
            "end_date": "",
            "cashier_id": "",
            "reset_rekap_msg": True,    # → untuk SweetAlert2
        }

    # =====================================================
    # FILTER NORMAL
    # =====================================================
    if start_date:
        sales = sales.filter(date_time__date__gte=start_date)
    if end_date:
        sales = sales.filter(date_time__date__lte=end_date)
    if cashier_id:
        sales = sales.filter(cashier_id=cashier_id)

    total_revenue = sales.aggregate(total=Sum("total"))["total"] or 0
    total_transactions = sales.count()

    # =====================================================
    # REKAP PER HARI
    # =====================================================
    per_day = (
        sales.annotate(day=TruncDate("date_time"))
        .values("day")
        .annotate(
            total=Sum("total"),
            count=Count("id")
        )
        .order_by("day")
    )

    rekap = []
    for row in per_day:
        day = row["day"]

        top_item = (
            SaleItem.objects.filter(sale__date_time__date=day)
            .values("product__name", "product__category__name")
            .annotate(qty=Sum("qty"))
            .order_by("-qty")
            .first()
        )

        rekap.append({
            "day": day,
            "count": row["count"],
            "total": row["total"],
            "product": top_item["product__name"] if top_item else "-",
            "category": top_item["product__category__name"] if top_item else "-",
        })

    return {
        "sales": sales,
        "rekap": rekap,
        "total_revenue": total_revenue,
        "total_transactions": total_transactions,
        "start_date": start_date or "",
        "end_date": end_date or "",
        "cashier_id": cashier_id or "",
    }







def sales_report(request):
    data = get_filtered_sales(request)
    cashiers = User.objects.filter(is_active=True).order_by("username")

    context = {
        "sales": data["sales"],
        "total_revenue": data["total_revenue"],
        "total_transactions": data["total_transactions"],
        "rekap": data["rekap"],
        "start_date": data["start_date"],
        "end_date": data["end_date"],
        "cashier_id": data["cashier_id"],
        "cashiers": cashiers,
    }

    return render(request, "kasir/sales_report.html", context)


def sales_report_export_excel(request):
    data = get_filtered_sales(request)
    sales = data["sales"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Penjualan"

    ws["A1"] = "Laporan Penjualan"
    ws["A2"] = f"Periode: {data['start_date']} s/d {data['end_date']}"
    ws.append([])
    ws.append(["Tanggal", "Invoice", "Kasir", "Total"])

    for sale in sales:
        ws.append(
            [
                sale.date_time.strftime("%Y-%m-%d %H:%M"),
                sale.invoice_no,
                sale.cashier.username if sale.cashier else "",
                float(sale.total),
            ]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="laporan_penjualan.xlsx"'
    wb.save(response)
    return response


def sales_report_export_pdf(request):
    data = get_filtered_sales(request)
    sales = data["sales"]

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="laporan_penjualan.pdf"'

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    y = height - 50
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, "Laporan Penjualan")
    y -= 20
    p.setFont("Helvetica", 10)
    p.drawString(50, y, f"Periode: {data['start_date']} s/d {data['end_date']}")
    y -= 30

    p.setFont("Helvetica-Bold", 9)
    p.drawString(50, y, "Tanggal")
    p.drawString(150, y, "Invoice")
    p.drawString(280, y, "Kasir")
    p.drawString(400, y, "Total")
    y -= 15
    p.setFont("Helvetica", 9)

    for sale in sales:
        if y < 50:
            p.showPage()
            y = height - 50
            p.setFont("Helvetica", 9)

        p.drawString(50, y, sale.date_time.strftime("%Y-%m-%d %H:%M"))
        p.drawString(150, y, sale.invoice_no)
        p.drawString(
            280, y, sale.cashier.username if sale.cashier else ""
        )
        p.drawRightString(480, y, f"{sale.total}")
        y -= 15

    p.showPage()
    p.save()
    return response


def struk_view(request, invoice_no):
    sale = Sale.objects.get(invoice_no=invoice_no)
    return render(request, "kasir/struk.html", {"sale": sale})


# ========== RESET DETAIL (hapus semua transaksi) ==========
def reset_detail(request):
    if request.method == "POST":
        SaleItem.objects.all().delete()
        Sale.objects.all().delete()
        messages.success(request, "Semua detail transaksi berhasil dihapus permanen!")
        return redirect("kasir:sales_report")



# ========== RESET REKAP (hapus semua rekap harian) ==========
def reset_rekap(request):
    if request.method == "POST":
        SaleItem.objects.all().delete()
        Sale.objects.all().delete()
        messages.success(request, "Semua rekap harian berhasil dihapus permanen!")
        return redirect("kasir:sales_report")

def export_detail_excel(request):
    sales = Sale.objects.all().order_by("-date_time")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detail Transaksi"

    # Header
    headers = [
        "Tanggal", "Invoice", "Kasir", "Metode",
        "Total", "Barang", "Kategori", "Jumlah"
    ]
    ws.append(headers)

    # Data
    for sale in sales:
        for item in sale.items.all():
            ws.append([
                sale.date_time.strftime("%Y-%m-%d %H:%M"),
                sale.invoice_no,
                sale.cashier.username,
                sale.payment_method,
                float(sale.total),
                item.product.name,
                item.product.category.name if item.product.category else "-",
                item.qty,
            ])

    # Auto width
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(column)].width = max_length + 2

    # Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="detail_transaksi.xlsx"'
    wb.save(response)
    return response

def export_detail_pdf(request):
    sales = Sale.objects.all().order_by("-date_time")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="detail_transaksi.pdf"'

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    y = height - 50
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, "Detail Transaksi")
    y -= 30

    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, y, "Tanggal")
    p.drawString(150, y, "Invoice")
    p.drawString(240, y, "Kasir")
    p.drawString(300, y, "Metode")
    p.drawString(360, y, "Barang")
    p.drawString(460, y, "Qty")
    y -= 20
    p.setFont("Helvetica", 9)

    for sale in sales:
        for item in sale.items.all():
            if y < 40:
                p.showPage()
                y = height - 50

            p.drawString(50, y, sale.date_time.strftime("%Y-%m-%d %H:%M"))
            p.drawString(150, y, sale.invoice_no)
            p.drawString(240, y, sale.cashier.username)
            p.drawString(300, y, sale.payment_method)
            p.drawString(360, y, item.product.name)
            p.drawString(460, y, str(item.qty))

            y -= 15

    p.showPage()
    p.save()
    return response

def logout_view(request):
    logout(request)
    return redirect("/accounts/login/")


